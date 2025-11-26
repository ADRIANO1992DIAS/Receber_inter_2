import base64
import calendar
import csv
import hashlib
import datetime as dt
import io
import unicodedata
import zipfile
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional, List, Set, Dict, Any, Tuple

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import FileResponse, HttpResponse
from django.db import transaction
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce, ExtractDay
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.views.decorators.http import require_POST, require_http_methods
from django.utils import timezone
from django.utils.text import slugify
from django.urls import reverse
from django.conf import settings

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import Cliente, Boleto, ConciliacaoLancamento, ConciliacaoAlias, WhatsappConfig, InterConfig
from .forms import (
    SelecionarClientesForm,
    ClienteForm,
    BoletoForm,
    ClienteImportForm,
    ConciliacaoUploadForm,
    ConciliacaoLinkForm,
    WhatsappMensagemForm,
    InterConfigForm,
)
from .services.inter_service import InterService
from .services.whatsapp_service import dispatch_boleto_via_whatsapp, format_whatsapp_phone


MESES_CHOICES = [
    (1, "Janeiro"),
    (2, "Fevereiro"),
    (3, "Marco"),
    (4, "Abril"),
    (5, "Maio"),
    (6, "Junho"),
    (7, "Julho"),
    (8, "Agosto"),
    (9, "Setembro"),
    (10, "Outubro"),
    (11, "Novembro"),
    (12, "Dezembro"),
]

CLIENTE_IMPORT_HEADER_ALIASES: Dict[str, str] = {
    "nome": "nome",
    "cliente": "nome",
    "razaosocial": "nome",
    "cpfcnpj": "cpfCnpj",
    "cpf": "cpfCnpj",
    "cnpj": "cpfCnpj",
    "documento": "cpfCnpj",
    "valornominal": "valorNominal",
    "valor": "valorNominal",
    "valorbruto": "valorNominal",
    "datavencimento": "dataVencimento",
    "vencimento": "dataVencimento",
    "diavencimento": "dataVencimento",
    "diadovencimento": "dataVencimento",
    "dia": "dataVencimento",
    "email": "email",
    "ddd": "ddd",
    "telefone": "telefone",
    "celular": "telefone",
    "endereco": "endereco",
    "logradouro": "endereco",
    "numero": "numero",
    "complemento": "complemento",
    "bairro": "bairro",
    "cidade": "cidade",
    "municipio": "cidade",
    "uf": "uf",
    "estado": "uf",
    "cep": "cep",
}

CLIENTE_IMPORT_REQUIRED = {"nome", "cpfCnpj", "valorNominal", "dataVencimento"}
DEFAULT_BOLETO_DDD = "85"
DEFAULT_BOLETO_TELEFONE = "985134478"


def _normalizar_header(valor: Optional[str]) -> str:
    if valor is None:
        return ""
    texto = unicodedata.normalize("NFKD", str(valor).strip().lower())
    return "".join(ch for ch in texto if ch.isalnum())


def _texto_limpo(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, str):
        return valor.strip()
    return str(valor).strip()


def _apenas_digitos(valor: str) -> str:
    return "".join(ch for ch in valor if ch.isdigit())


def _parse_decimal(valor) -> Decimal:
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        raise ValueError("Valor nominal ausente.")

    if isinstance(valor, Decimal):
        decimal_valor = valor
    elif isinstance(valor, (int, float)):
        decimal_valor = Decimal(str(valor))
    else:
        texto = str(valor)
        texto = texto.replace("R$", "").replace(" ", "")
        if "," in texto and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(",", ".")
        decimal_valor = Decimal(texto)

    return decimal_valor.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _parse_dia_vencimento(valor) -> int:
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        raise ValueError("Campo dataVencimento ausente.")

    if isinstance(valor, dt.date):
        dia = valor.day
    elif isinstance(valor, (int, float, Decimal)):
        dia = int(valor)
    else:
        texto = str(valor).strip()
        if not texto:
            raise ValueError("Campo dataVencimento vazio.")
        dia = int(float(texto.replace(",", ".")))

    if not 1 <= dia <= 31:
        raise ValueError("Campo dataVencimento deve estar entre 1 e 31.")
    return dia


def _arquivo_pdf_nome(boleto: Boleto) -> str:
    competencia = f"{boleto.competencia_mes:02d}-{boleto.competencia_ano}"
    base = f"{boleto.cliente.nome}-{competencia}-{boleto.id}"
    slug = slugify(base)
    if not slug:
        slug = f"boleto-{boleto.id}"
    return f"{slug}.pdf"


def _pdf_existe_localmente(boleto: Boleto) -> bool:
    """
    Confirma se o arquivo referenciado pelo campo FileField realmente existe no storage.
    """
    if not boleto.pdf or not boleto.pdf.name:
        return False
    try:
        return boleto.pdf.storage.exists(boleto.pdf.name)
    except OSError:
        return False


def _preparar_boleto_para_reemissao(boleto: Boleto) -> None:
    """
    Limpa dados sensÃ­veis de um boleto cancelado para permitir nova emissÃ£o.
    """
    if boleto.pdf:
        try:
            boleto.pdf.delete(save=False)
        except FileNotFoundError:
            pass
    boleto.pdf = None
    boleto.nosso_numero = ""
    boleto.linha_digitavel = ""
    boleto.codigo_barras = ""
    boleto.tx_id = ""
    boleto.codigo_solicitacao = ""
    boleto.status = Boleto.STATUS_NOVO
    boleto.erro_msg = ""
    boleto.data_pagamento = None
    boleto.forma_pagamento = ""
    boleto.whatsapp_status = Boleto.WHATSAPP_STATUS_PENDENTE
    boleto.whatsapp_status_detail = ""
    boleto.whatsapp_status_updated_at = None


def _atualizar_codigo_barras_via_inter(inter: InterService, boleto: Boleto) -> None:
    if boleto.codigo_barras:
        return

    identificadores = [
        (boleto.nosso_numero, "nosso_numero"),
        (boleto.codigo_solicitacao, "codigo_solicitacao"),
        (boleto.tx_id, "tx_id"),
    ]
    for ident, campo in identificadores:
        if not ident:
            continue
        try:
            detalhe = inter.recuperar_cobranca_detalhada(ident, campo=campo)
        except Exception:
            continue
        if not detalhe:
            continue

        campos_atualizados: List[str] = []
        codigo = detalhe.get("codigoBarras")
        if codigo and codigo != boleto.codigo_barras:
            boleto.codigo_barras = codigo
            campos_atualizados.append("codigo_barras")
        linha_digitavel = detalhe.get("linhaDigitavel")
        if linha_digitavel and not boleto.linha_digitavel:
            boleto.linha_digitavel = linha_digitavel
            campos_atualizados.append("linha_digitavel")

        if campos_atualizados:
            boleto.save(update_fields=campos_atualizados)
            return


def _buscar_pdf_bytes(inter: InterService, boleto: Boleto) -> Optional[bytes]:
    if _pdf_existe_localmente(boleto):
        try:
            with boleto.pdf.open("rb") as stream:
                return stream.read()
        except FileNotFoundError:
            # Arquivo foi removido do disco; forÃ§a re-download via API.
            pass

    identificadores = [
        (boleto.nosso_numero, "nosso_numero"),
        (boleto.codigo_solicitacao, "codigo_solicitacao"),
    ]
    for ident, campo in identificadores:
        if not ident:
            continue
        pdf_bytes = inter.baixar_pdf(ident, campo=campo)
        if pdf_bytes:
            if isinstance(pdf_bytes, str):
                pdf_bytes = base64.b64decode(pdf_bytes)
            return pdf_bytes
    return None


def _parse_inter_date(valor: Optional[str]) -> Optional[dt.date]:
    if not valor:
        return None
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    texto = str(valor).strip()
    if not texto:
        return None
    texto = texto.replace("Z", "")
    try:
        return dt.datetime.fromisoformat(texto).date()
    except ValueError:
        pass
    if "T" in texto:
        try:
            return dt.datetime.fromisoformat(texto.split("T")[0]).date()
        except ValueError:
            pass
    for formato in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _interpretar_status_cobranca(payload: Dict[str, Any]) -> Dict[str, Optional[dt.date]]:
    if not isinstance(payload, dict):
        return {"status": None, "data_pagamento": None}

    blocos: List[Dict[str, Any]] = [payload]
    for chave in ("cobranca", "boleto", "pix", "detalhes", "dadosPagamento"):
        valor = payload.get(chave)
        if isinstance(valor, dict):
            blocos.append(valor)

    pagamentos_coletados: List[Dict[str, Any]] = []
    for bloco in blocos:
        tot_pag = bloco.get("pagamentos") or bloco.get("listaPagamentos")
        if isinstance(tot_pag, list):
            pagamentos_coletados.extend([p for p in tot_pag if isinstance(p, dict)])

    status_candidatos: List[str] = []
    for bloco in blocos:
        for chave in (
            "situacao",
            "status",
            "situacaoAtual",
            "situacaoAtualCobranca",
            "statusCobranca",
            "situacaoBoleto",
            "statusBoleto",
        ):
            valor = bloco.get(chave)
            if valor:
                status_candidatos.append(str(valor))

    def _normalize(texto: str) -> str:
        return (
            unicodedata.normalize("NFKD", texto or "")
            .encode("ASCII", "ignore")
            .decode()
            .upper()
            .replace(" ", "")
        )

    novo_status: Optional[str] = None
    status_atraso_tokens = {"ATRASADO", "ATRASADA", "VENCIDO", "VENCIDA", "VENCID", "EMATRASO"}
    status_cancelamento_tokens = {"CANCEL", "BAIXA", "EXPIR", "DEVOL"}
    status_pago_tokens = {"PAGO", "LIQUID", "BAIXADO", "RECEBIDO", "LIQUIDADO"}
    status_aberto_tokens = {"EMABERTO", "ABERTO", "EMISSAO", "EMITIDO", "EMITIDA"}

    for status_bruto in status_candidatos:
        status_normalizado = _normalize(status_bruto)
        if any(chave in status_normalizado for chave in status_pago_tokens):
            novo_status = Boleto.STATUS_PAGO
            break
        if any(chave in status_normalizado for chave in status_cancelamento_tokens):
            novo_status = Boleto.STATUS_CANCELADO
            break
        if any(chave in status_normalizado for chave in status_atraso_tokens):
            novo_status = Boleto.STATUS_ATRASADO
            continue
        if not novo_status and status_normalizado:
            if any(chave in status_normalizado for chave in status_aberto_tokens):
                novo_status = Boleto.STATUS_EMITIDO
            else:
                novo_status = Boleto.STATUS_EMITIDO

    def _valor_para_decimal(valor: Any) -> Optional[Decimal]:
        if valor in (None, "", "None"):
            return None
        try:
            return Decimal(str(valor)).quantize(Decimal("0.01"))
        except (InvalidOperation, TypeError, ValueError):
            return None

    valores_para_checar = []
    for bloco in blocos:
        valores_para_checar.extend(
            bloco.get(chave)
            for chave in ("valorPago", "valorRecebido", "valorLiquidado", "valorQuitado")
            if chave in bloco
        )
    for pagamento in pagamentos_coletados:
        valores_para_checar.extend(
            pagamento.get(chave)
            for chave in ("valorPago", "valor", "valorLiquidado")
            if chave in pagamento
        )

    for valor_bruto in valores_para_checar:
        valor_convertido = _valor_para_decimal(valor_bruto)
        if valor_convertido and valor_convertido > Decimal("0"):
            novo_status = Boleto.STATUS_PAGO
            break

    if not novo_status:
        for pagamento in pagamentos_coletados:
            situacao_pagamento = pagamento.get("situacao") or pagamento.get("status")
            if situacao_pagamento:
                situacao_normalizada = _normalize(str(situacao_pagamento))
                if any(chave in situacao_normalizada for chave in status_pago_tokens):
                    novo_status = Boleto.STATUS_PAGO
                    break
                if any(chave in situacao_normalizada for chave in status_cancelamento_tokens):
                    novo_status = Boleto.STATUS_CANCELADO
                    break
                if any(chave in situacao_normalizada for chave in status_atraso_tokens):
                    novo_status = Boleto.STATUS_ATRASADO
                    break

    data_pagamento_bruta: Optional[str] = None
    for bloco in blocos:
        data_pagamento_bruta = bloco.get("dataPagamento") or bloco.get("dataPagto") or bloco.get("dataPagamentoBoleto")
        if data_pagamento_bruta:
            break

    if not data_pagamento_bruta and pagamentos_coletados:
        for pagamento in pagamentos_coletados:
            data_pagamento_bruta = (
                pagamento.get("dataPagamento")
                or pagamento.get("dataHoraPagamento")
                or pagamento.get("dataHora")
                or pagamento.get("data")
            )
            if data_pagamento_bruta:
                break

    data_pagamento = _parse_inter_date(data_pagamento_bruta)

    return {
        "status": novo_status,
        "data_pagamento": data_pagamento,
    }


def home(request):
    # Agora a raiz (/) redireciona para a lista de clientes
    return redirect("clientes_list")


@login_required
def clientes_list(request):
    clientes_qs = Cliente.objects.all()

    nome_param = request.GET.get("nome", "").strip()
    dia_param = request.GET.get("dia_vencimento", "").strip()
    valor_min_param = request.GET.get("valor_min", "").strip()
    valor_max_param = request.GET.get("valor_max", "").strip()
    if "status" in request.GET:
        status_param_raw = request.GET.get("status", "")
    else:
        status_param_raw = Boleto.STATUS_EMITIDO
    status_param = (status_param_raw or "").strip().lower()

    if nome_param:
        clientes_qs = clientes_qs.filter(nome__icontains=nome_param)

    dia_vencimento = None
    if dia_param:
        try:
            dia_vencimento = int(dia_param)
        except ValueError:
            dia_param = ""
        else:
            if 1 <= dia_vencimento <= 31:
                clientes_qs = clientes_qs.filter(dataVencimento=dia_vencimento)
            else:
                dia_param = ""

    valor_min = None
    if valor_min_param:
        texto = valor_min_param.replace("R$", "").replace(" ", "")
        if texto.count(",") == 1 and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", ".")
        try:
            valor_min = Decimal(texto)
        except (InvalidOperation, ValueError):
            valor_min_param = ""
        else:
            clientes_qs = clientes_qs.filter(valorNominal__gte=valor_min)

    valor_max = None
    if valor_max_param:
        texto = valor_max_param.replace("R$", "").replace(" ", "")
        if texto.count(",") == 1 and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", ".")
        try:
            valor_max = Decimal(texto)
        except (InvalidOperation, ValueError):
            valor_max_param = ""
        else:
            clientes_qs = clientes_qs.filter(valorNominal__lte=valor_max)

    if status_param == "ativos":
        clientes_qs = clientes_qs.filter(ativo=True)
    elif status_param == "inativos":
        clientes_qs = clientes_qs.filter(ativo=False)
    else:
        status_param = ""

    clientes = clientes_qs.order_by("nome")
    filtros_aplicados = {
        "nome": nome_param,
        "dia_vencimento": dia_param,
        "valor_min": valor_min_param,
        "valor_max": valor_max_param,
        "status": status_param,
    }
    filtros_ativos = any(filtros_aplicados.values())
    return render(
        request,
        "billing/clientes_list.html",
        {"clientes": clientes, "filtros": filtros_aplicados, "filtros_ativos": filtros_ativos},
    )


@login_required
def dashboard(request):
    boletos_qs = Boleto.objects.select_related("cliente").all()

    mes_params = [valor.strip() for valor in request.GET.getlist("mes")]
    ano_params = [valor.strip() for valor in request.GET.getlist("ano")]
    dia_params = [valor.strip() for valor in request.GET.getlist("dia")]

    hoje = timezone.localdate()

    meses_selecionados: List[str] = []
    anos_selecionados: List[str] = []
    dias_selecionados: List[str] = []

    if "mes" not in request.GET:
        if hoje:
            boletos_qs = boletos_qs.filter(competencia_mes=hoje.month)
            meses_selecionados = [str(hoje.month)]
    elif "" in mes_params:
        meses_selecionados = [""]
    else:
        meses_validos: List[int] = []
        for valor in mes_params:
            if not valor:
                continue
            try:
                mes_valor = int(valor)
            except ValueError:
                continue
            if 1 <= mes_valor <= 12 and mes_valor not in meses_validos:
                meses_validos.append(mes_valor)
                meses_selecionados.append(str(mes_valor))
        if meses_validos:
            boletos_qs = boletos_qs.filter(competencia_mes__in=meses_validos)

    if "ano" not in request.GET:
        if hoje:
            boletos_qs = boletos_qs.filter(competencia_ano=hoje.year)
            anos_selecionados = [str(hoje.year)]
    elif "" in ano_params:
        anos_selecionados = [""]
    else:
        anos_validos: List[int] = []
        for valor in ano_params:
            if not valor:
                continue
            try:
                ano_valor = int(valor)
            except ValueError:
                continue
            if ano_valor not in anos_validos:
                anos_validos.append(ano_valor)
                anos_selecionados.append(str(ano_valor))
        if anos_validos:
            boletos_qs = boletos_qs.filter(competencia_ano__in=anos_validos)

    if "" in dia_params:
        dias_selecionados = [""]
    else:
        dias_validos: List[int] = []
        for valor in dia_params:
            if not valor:
                continue
            try:
                dia_valor = int(valor)
            except ValueError:
                continue
            if 1 <= dia_valor <= 31 and dia_valor not in dias_validos:
                dias_validos.append(dia_valor)
                dias_selecionados.append(str(dia_valor))
        if dias_validos:
            boletos_qs = boletos_qs.filter(data_vencimento__day__in=dias_validos)

    total_gerados = boletos_qs.count()
    total_recebidos = boletos_qs.filter(status=Boleto.STATUS_PAGO).count()
    total_cancelados = boletos_qs.filter(status=Boleto.STATUS_CANCELADO).count()
    total_valor_gerado = boletos_qs.aggregate(total=Coalesce(Sum("valor"), Decimal("0")))["total"]
    boletos_recebidos = boletos_qs.filter(status=Boleto.STATUS_PAGO)
    total_valor_recebido = boletos_recebidos.aggregate(total=Coalesce(Sum("valor"), Decimal("0")))["total"]
    total_valor_cancelado = boletos_qs.filter(status=Boleto.STATUS_CANCELADO).aggregate(total=Coalesce(Sum("valor"), Decimal("0")))["total"]

    boletos_pix = boletos_recebidos.filter(forma_pagamento="pix")
    boletos_dinheiro = boletos_recebidos.filter(forma_pagamento="dinheiro")
    total_pix = boletos_pix.count()
    total_dinheiro = boletos_dinheiro.count()
    valor_pix = boletos_pix.aggregate(total=Coalesce(Sum("valor"), Decimal("0")))["total"]
    valor_dinheiro = boletos_dinheiro.aggregate(total=Coalesce(Sum("valor"), Decimal("0")))["total"]
    total_pix_dinheiro = total_pix + total_dinheiro
    valor_pix_dinheiro = (valor_pix or Decimal("0")) + (valor_dinheiro or Decimal("0"))

    hoje = hoje or timezone.localdate()
    boletos_em_aberto = boletos_qs.filter(
        status__in=[
            Boleto.STATUS_EMITIDO,
            Boleto.STATUS_NOVO,
            Boleto.STATUS_ATRASADO,
        ]
    )
    boletos_atrasados = boletos_qs.filter(
        Q(status=Boleto.STATUS_ATRASADO)
        | (
            Q(
                status__in=[
                    Boleto.STATUS_EMITIDO,
                    Boleto.STATUS_NOVO,
                ]
            )
            & Q(data_vencimento__lt=hoje)
        )
    )
    total_em_atraso = boletos_atrasados.count()
    valor_em_atraso = boletos_atrasados.aggregate(total=Coalesce(Sum("valor"), Decimal("0")))["total"]

    boletos_a_receber = boletos_qs.filter(
        Q(
            status__in=[
                Boleto.STATUS_EMITIDO,
                Boleto.STATUS_NOVO,
            ]
        )
        & (Q(data_vencimento__gte=hoje) | Q(data_vencimento__isnull=True))
    )
    total_a_receber = boletos_a_receber.count()
    valor_a_receber = boletos_a_receber.aggregate(total=Coalesce(Sum("valor"), Decimal("0")))["total"]

    anos_disponiveis = list(
        Boleto.objects.order_by("-competencia_ano")
        .values_list("competencia_ano", flat=True)
        .distinct()
    )
    if hoje and hoje.year not in anos_disponiveis:
        anos_disponiveis.append(hoje.year)
    anos_disponiveis = sorted({int(ano) for ano in anos_disponiveis}, reverse=True)

    dias_disponiveis = (
        Boleto.objects.annotate(dia=ExtractDay("data_vencimento"))
        .values_list("dia", flat=True)
        .order_by("dia")
        .distinct()
    )

    meses_contexto = [{"value": "", "label": "Todos"}] + [
        {"value": str(valor), "label": nome} for valor, nome in MESES_CHOICES
    ]

    dias_contexto = [{"value": "", "label": "Todos"}] + [
        {"value": str(dia), "label": f"Dia {int(dia):02d}"}
        for dia in dias_disponiveis
        if dia is not None
    ]

    def _resumo_selecao(opcoes, selecionados):
        if "" in selecionados or not selecionados:
            return "Todos"
        valores_validos = [valor for valor in selecionados if valor]
        if not valores_validos:
            return "Todos"
        if len(valores_validos) > 1:
            return "Diversos"
        alvo = valores_validos[0]
        for opcao in opcoes:
            if isinstance(opcao, dict):
                if str(opcao.get("value", "")) == alvo:
                    return str(opcao.get("label", alvo))
            else:
                if str(opcao) == alvo:
                    return str(opcao)
        return str(alvo)

    anos_contexto = [{"value": "", "label": "Todos"}] + [
        {"value": str(ano), "label": str(ano)} for ano in anos_disponiveis
    ]

    ultimos_boletos = (
        boletos_qs.order_by("-criado_em")[:5]
        if total_gerados
        else []
    )

    context = {
        "total_gerados": total_gerados,
        "total_recebidos": total_recebidos,
        "total_cancelados": total_cancelados,
        "valor_gerado": total_valor_gerado,
        "valor_recebido": total_valor_recebido,
        "valor_cancelado": total_valor_cancelado,
        "total_pix_dinheiro": total_pix_dinheiro,
        "valor_pix_dinheiro": valor_pix_dinheiro,
        "total_em_atraso": total_em_atraso,
        "valor_em_atraso": valor_em_atraso,
        "total_a_receber": total_a_receber,
        "valor_a_receber": valor_a_receber,
        "meses": meses_contexto,
        "anos": [str(ano) for ano in anos_disponiveis],
        "dias": dias_contexto,
        "meses_selecionados": meses_selecionados,
        "anos_selecionados": anos_selecionados,
        "dias_selecionados": dias_selecionados,
        "resumo_meses": _resumo_selecao(meses_contexto, meses_selecionados),
        "resumo_anos": _resumo_selecao(anos_contexto, anos_selecionados),
        "resumo_dias": _resumo_selecao(dias_contexto, dias_selecionados),
        "ultimos_boletos": ultimos_boletos,
    }
    return render(request, "billing/dashboard.html", context)


def _hash_conciliacao(data: dt.date, descricao: str, valor: Decimal) -> str:
    descricao_normalizada = (
        unicodedata.normalize("NFKD", descricao or "")
        .encode("ASCII", "ignore")
        .decode()
        .strip()
        .lower()
    )
    valor_formatado = format(valor.quantize(Decimal("0.01")), ".2f")
    base = f"{data.isoformat()}|{descricao_normalizada}|{valor_formatado}"
    return hashlib.sha1(base.encode("utf-8")).hexdigest()


def _normalizar_texto_para_match(texto: str) -> str:
    if not texto:
        return ""
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode()
    return " ".join(texto.lower().split())


def _carregar_conciliacao_csv(arquivo) -> Tuple[List["ConciliacaoLancamento"], int]:
    bruto = arquivo.read()
    try:
        texto = bruto.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = bruto.decode("latin-1")
    finally:
        arquivo.seek(0)

    leitor = csv.reader(io.StringIO(texto), delimiter=";")
    registros: List[ConciliacaoLancamento] = []
    cabecalho_encontrado = False
    auto_baixas = 0

    for linha in leitor:
        colunas = [col.strip() for col in linha]
        if not any(colunas):
            continue
        if not cabecalho_encontrado:
            colunas_normalizadas = [
                unicodedata.normalize("NFKD", coluna or "").lower() for coluna in colunas
            ]
            if len(colunas_normalizadas) >= 4 and "data" in colunas_normalizadas[0] and "valor" in colunas_normalizadas[3]:
                cabecalho_encontrado = True
            continue
        if len(colunas) < 4:
            continue

        data_txt, _, historico_txt, valor_txt, *restante = colunas
        descricao_txt = (colunas[2] or historico_txt or "").strip()
        descricao_chave = _normalizar_texto_para_match(descricao_txt)
        if not data_txt or not valor_txt:
            continue
        try:
            data = dt.datetime.strptime(data_txt, "%d/%m/%Y").date()
        except ValueError:
            continue
        valor_normalizado = (
            valor_txt.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
        )
        try:
            valor = Decimal(valor_normalizado).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except (InvalidOperation, ValueError):
            continue

        hash_identificador = _hash_conciliacao(data, descricao_txt, valor)
        lancamento, created = ConciliacaoLancamento.objects.get_or_create(
            hash_identificador=hash_identificador,
            defaults={
                "data": data,
                "descricao": descricao_txt,
                "descricao_chave": descricao_chave,
                "valor": valor,
            },
        )

        campos_para_atualizar: List[str] = []
        if not created:
            if lancamento.data != data:
                lancamento.data = data
                campos_para_atualizar.append("data")
            if lancamento.descricao != descricao_txt:
                lancamento.descricao = descricao_txt
                campos_para_atualizar.append("descricao")
            if lancamento.descricao_chave != descricao_chave:
                lancamento.descricao_chave = descricao_chave
                campos_para_atualizar.append("descricao_chave")
            if lancamento.valor != valor:
                lancamento.valor = valor
                campos_para_atualizar.append("valor")
            if campos_para_atualizar:
                lancamento.save(update_fields=campos_para_atualizar + ["atualizado_em"])

        registros.append(lancamento)

        alias_cliente: Optional[ConciliacaoAlias] = None
        if descricao_chave:
            alias_cliente = (
                ConciliacaoAlias.objects.select_related("cliente")
                .filter(descricao_chave=descricao_chave)
                .first()
            )

        boleto = lancamento.boleto
        if boleto and boleto.status in (Boleto.STATUS_EMITIDO, Boleto.STATUS_ATRASADO):
            _registrar_pagamento_manual(boleto, forma_pagamento="pix", data_pagamento=lancamento.data)
            if alias_cliente is None and descricao_chave:
                ConciliacaoAlias.objects.get_or_create(
                    descricao_chave=descricao_chave,
                    defaults={"cliente": boleto.cliente},
                )
            auto_baixas += 1
            continue

        if alias_cliente and not lancamento.boleto_id:
            boleto_auto = (
                Boleto.objects.filter(
                    cliente=alias_cliente.cliente,
                    status__in=[Boleto.STATUS_EMITIDO, Boleto.STATUS_ATRASADO],
                    valor=lancamento.valor,
                )
                .order_by("data_vencimento", "id")
                .first()
            )
            if boleto_auto:
                lancamento.boleto = boleto_auto
                lancamento.save(update_fields=["boleto", "atualizado_em"])
                _registrar_pagamento_manual(boleto_auto, forma_pagamento="pix", data_pagamento=lancamento.data)
                auto_baixas += 1

    if not cabecalho_encontrado:
        raise ValueError("Cabecalho 'Data;Descricao;Valor' nao encontrado no arquivo CSV.")
    if not registros:
        raise ValueError("Nenhum registro valido foi encontrado no arquivo.")
    return registros, auto_baixas


@login_required
def conciliacao(request):
    pendentes_only = request.GET.get("pendentes") == "1"
    upload_form = ConciliacaoUploadForm()
    novos_ids: Set[int] = set()
    auto_baixas = 0

    if request.method == "POST":
        acao = request.POST.get("acao") or "upload"
        if acao == "apagar_pendentes":
            removidos, _ = ConciliacaoLancamento.objects.filter(boleto__isnull=True).delete()
            if removidos:
                messages.success(request, f"{removidos} lancamento(s) sem vinculo removido(s).")
            else:
                messages.info(request, "Nenhum lancamento pendente para remover.")
            redirect_url = reverse("conciliacao")
            if pendentes_only:
                redirect_url += "?pendentes=1"
            return redirect(redirect_url)
        if acao == "vincular":
            link_form = ConciliacaoLinkForm(request.POST)
            if link_form.is_valid():
                lancamento = link_form.cleaned_data["lancamento"]
                boleto = link_form.cleaned_data["boleto"]
                descricao_chave = lancamento.descricao_chave or _normalizar_texto_para_match(lancamento.descricao)
                update_fields = ["boleto", "atualizado_em"]
                if descricao_chave and lancamento.descricao_chave != descricao_chave:
                    lancamento.descricao_chave = descricao_chave
                    update_fields.append("descricao_chave")
                lancamento.boleto = boleto
                lancamento.save(update_fields=update_fields)
                if descricao_chave:
                    alias, created = ConciliacaoAlias.objects.get_or_create(
                        descricao_chave=descricao_chave,
                        defaults={"cliente": boleto.cliente},
                    )
                    if not created and alias.cliente_id != boleto.cliente_id:
                        alias.cliente = boleto.cliente
                        alias.save(update_fields=["cliente", "atualizado_em"])
                _registrar_pagamento_manual(boleto, forma_pagamento="pix", data_pagamento=lancamento.data)
                messages.success(
                    request,
                    f"Lancamento de {lancamento.descricao} vinculado ao boleto #{boleto.id} e marcado como pago.",
                )
                redirect_url = reverse("conciliacao")
                if pendentes_only:
                    redirect_url += "?pendentes=1"
                return redirect(redirect_url)
            for erro in link_form.errors.get("__all__", []):
                messages.error(request, erro)
            for campo, erros in link_form.errors.items():
                if campo == "__all__":
                    continue
                for erro in erros:
                    messages.error(request, erro)
        else:
            upload_form = ConciliacaoUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                arquivo = upload_form.cleaned_data["arquivo"]
                try:
                    registros_importados, auto_baixas = _carregar_conciliacao_csv(arquivo)
                except ValueError as exc:
                    upload_form.add_error(None, str(exc))
                else:
                    novos_ids = {reg.id for reg in registros_importados}
                    total = len(registros_importados)
                    if total:
                        mensagem = f"{total} registro(s) processado(s) do extrato."
                        if auto_baixas:
                            mensagem += f" {auto_baixas} boleto(s) conciliado(s) automaticamente."
                        messages.success(request, mensagem)
                    else:
                        messages.info(request, "Arquivo processado, mas nenhum novo lancamento foi encontrado.")
            else:
                mensagens = upload_form.errors.get("__all__", [])
                for erro in mensagens:
                    messages.error(request, erro)

    boletos_elegiveis = list(
        Boleto.objects.filter(
            status__in=[Boleto.STATUS_EMITIDO, Boleto.STATUS_ATRASADO]
        )
        .select_related("cliente")
        .order_by("cliente__nome", "competencia_ano", "competencia_mes")
    )

    registros_queryset = ConciliacaoLancamento.objects.select_related("boleto", "boleto__cliente")
    if pendentes_only:
        registros_queryset = registros_queryset.filter(boleto__isnull=True)
    registros_queryset = registros_queryset.order_by("-data", "-id")[:200]

    registros_contexto = []
    for lancamento in registros_queryset:
        descricao_normalizada = _normalizar_texto_para_match(lancamento.descricao)
        sugestoes: List[Dict[str, Any]] = []

        for boleto in boletos_elegiveis:
            valor_diff = abs(
                (boleto.valor - lancamento.valor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            )
            cliente_normalizado = _normalizar_texto_para_match(boleto.cliente.nome)
            similaridade = (
                SequenceMatcher(None, descricao_normalizada, cliente_normalizado).ratio()
                if descricao_normalizada and cliente_normalizado
                else 0.0
            )
            sugestoes.append(
                {
                    "boleto": boleto,
                    "valor_diff": valor_diff,
                    "valor_match": valor_diff == Decimal("0.00"),
                    "similaridade": similaridade,
                    "similaridade_percent": int(round(similaridade * 100)),
                }
            )

        sugestoes.sort(
            key=lambda item: (
                item["valor_diff"],
                -item["similaridade"],
                item["boleto"].id,
            )
        )
        sugestoes = sugestoes[:10]

        melhor_sugestao = sugestoes[0] if sugestoes else None
        melhor_similaridade = melhor_sugestao["similaridade"] if melhor_sugestao else 0.0

        registros_contexto.append(
            {
                "lancamento": lancamento,
                "sugestoes": sugestoes,
                "is_novo": lancamento.id in novos_ids,
                "melhor_sugestao": melhor_sugestao,
                "melhor_similaridade": melhor_similaridade,
            }
        )

    registros_contexto.sort(
        key=lambda item: (
            0 if item["lancamento"].boleto_id is None else 1,
            -item["melhor_similaridade"],
            -item["lancamento"].data.toordinal(),
            -item["lancamento"].id,
        )
    )

    pendentes_total = ConciliacaoLancamento.objects.filter(boleto__isnull=True).count()

    return render(
        request,
        "billing/conciliacao.html",
        {
            "upload_form": upload_form,
            "registros": registros_contexto,
            "boletos_disponiveis": boletos_elegiveis,
            "boletos_total": len(boletos_elegiveis),
            "auto_baixas": auto_baixas,
            "pendentes_total": pendentes_total,
            "pendentes_only": pendentes_only,
        },
    )


@login_required
@require_POST
def sincronizar_boletos(request):
    boletos = list(
        Boleto.objects.filter(
            status__in=[
                Boleto.STATUS_EMITIDO,
                Boleto.STATUS_NOVO,
                Boleto.STATUS_ERRO,
                Boleto.STATUS_ATRASADO,
            ]
        ).select_related("cliente")
    )
    if not boletos:
        messages.info(request, "Nenhum boleto pendente para sincronizar.")
        return redirect("boletos_list")

    try:
        inter = InterService()
    except Exception as exc:  # noqa: BLE001
        messages.error(request, f"Falha ao inicializar integracao com o Banco Inter: {exc}")
        return redirect("boletos_list")

    atualizados = 0
    contagem: Dict[str, int] = {
        Boleto.STATUS_PAGO: 0,
        Boleto.STATUS_CANCELADO: 0,
        Boleto.STATUS_EMITIDO: 0,
        Boleto.STATUS_ATRASADO: 0,
    }
    sem_detalhe = 0
    erros: List[str] = []

    for boleto in boletos:
        detalhe: Optional[Dict[str, Any]] = None
        for ident, campo in [
            (boleto.nosso_numero, "nosso_numero"),
            (boleto.codigo_solicitacao, "codigo_solicitacao"),
            (boleto.tx_id, "tx_id"),
        ]:
            if not ident:
                continue
            try:
                detalhe = inter.recuperar_cobranca_detalhada(ident, campo=campo)
            except Exception as exc:  # noqa: BLE001
                erros.append(f"Boleto {boleto.id} - {boleto.cliente.nome}: {exc}")
                detalhe = None
                break
            if detalhe:
                break

        if not detalhe:
            sem_detalhe += 1
            continue

        resultado = _interpretar_status_cobranca(detalhe)
        novo_status = resultado.get("status")
        data_pagamento = resultado.get("data_pagamento")

        if not novo_status:
            continue

        update_fields: Set[str] = set()

        if detalhe.get("nossoNumero") and detalhe["nossoNumero"] != boleto.nosso_numero:
            boleto.nosso_numero = detalhe["nossoNumero"]
            update_fields.add("nosso_numero")

        if detalhe.get("codigoSolicitacao") and detalhe["codigoSolicitacao"] != boleto.codigo_solicitacao:
            boleto.codigo_solicitacao = detalhe["codigoSolicitacao"]
            update_fields.add("codigo_solicitacao")

        if detalhe.get("linhaDigitavel") and detalhe["linhaDigitavel"] != boleto.linha_digitavel:
            boleto.linha_digitavel = detalhe["linhaDigitavel"]
            update_fields.add("linha_digitavel")

        if detalhe.get("codigoBarras") and detalhe["codigoBarras"] != boleto.codigo_barras:
            boleto.codigo_barras = detalhe["codigoBarras"]
            update_fields.add("codigo_barras")

        if detalhe.get("valorNominal"):
            try:
                valor_remote = Decimal(str(detalhe["valorNominal"]))
            except (InvalidOperation, TypeError, ValueError):
                valor_remote = None
            else:
                if valor_remote is not None and boleto.valor != valor_remote:
                    boleto.valor = valor_remote
                    update_fields.add("valor")

        if novo_status == Boleto.STATUS_PAGO:
            if data_pagamento and boleto.data_pagamento != data_pagamento:
                boleto.data_pagamento = data_pagamento
                update_fields.add("data_pagamento")
        else:
            if boleto.data_pagamento:
                boleto.data_pagamento = None
                update_fields.add("data_pagamento")
            if boleto.forma_pagamento:
                boleto.forma_pagamento = ""
                update_fields.add("forma_pagamento")

        if boleto.status != novo_status:
            boleto.status = novo_status
            update_fields.add("status")
            contagem[novo_status] = contagem.get(novo_status, 0) + 1

        if update_fields:
            boleto.save(update_fields=list(update_fields))
            atualizados += 1

    if atualizados:
        resumo_itens = []
        resumir_status = {
            Boleto.STATUS_PAGO: "recebidos",
            Boleto.STATUS_CANCELADO: "cancelados",
            Boleto.STATUS_ATRASADO: "atrasados",
            Boleto.STATUS_EMITIDO: "em aberto",
        }
        for status_codigo, descricao in resumir_status.items():
            quantidade = contagem.get(status_codigo)
            if quantidade:
                resumo_itens.append(f"{descricao}: {quantidade}")
        resumo = ", ".join(resumo_itens)
        mensagem = f"Sincronizacao concluida. {atualizados} boleto(s) atualizado(s)."
        if resumo:
            mensagem += f" ({resumo})"
        messages.success(request, mensagem)
    else:
        messages.info(request, "Sincronizacao concluida. Nenhum boleto precisava de atualizacao.")

    if sem_detalhe:
        messages.info(request, f"{sem_detalhe} boleto(s) nao foram encontrados ou ainda nao estao disponiveis na API.")
    if erros:
        mensagens = "; ".join(erros[:3])
        if len(erros) > 3:
            mensagens += f"; ... (+{len(erros) - 3} erro(s))"
        messages.warning(request, f"Algumas consultas falharam: {mensagens}")

    return redirect("boletos_list")


@login_required
def cliente_import(request):
    form = ClienteImportForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        arquivo = form.cleaned_data["arquivo"]
        arquivo.seek(0)

        try:
            workbook = load_workbook(arquivo, data_only=True)
        except InvalidFileException:
            form.add_error("arquivo", "O arquivo deve estar em formato Excel (.xlsx).")
        except Exception as exc:
            form.add_error("arquivo", f"NÃ£o foi possÃ­vel ler a planilha: {exc}")
        else:
            try:
                sheet = workbook.active
                header_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
                if not header_row:
                    form.add_error("arquivo", "A planilha precisa ter uma linha de cabeÃ§alho.")
                else:
                    header_map: Dict[str, int] = {}
                    for idx, header in enumerate(header_row):
                        campo = CLIENTE_IMPORT_HEADER_ALIASES.get(_normalizar_header(header))
                        if campo and campo not in header_map:
                            header_map[campo] = idx

                    campos_faltando = [campo for campo in CLIENTE_IMPORT_REQUIRED if campo not in header_map]
                    if campos_faltando:
                        cabecalhos = ", ".join(sorted(campos_faltando))
                        form.add_error("arquivo", f"CabeÃ§alhos obrigatÃ³rios ausentes: {cabecalhos}.")
                    else:
                        criados = atualizados = 0
                        erros: List[str] = []

                        for linha_idx, row in enumerate(
                            sheet.iter_rows(min_row=2, values_only=True),
                            start=2,
                        ):
                            if row is None:
                                continue

                            if all(
                                cell is None or (isinstance(cell, str) and not cell.strip())
                                for cell in row
                            ):
                                continue

                            dados = {
                                campo: (row[idx] if idx < len(row) else None)
                                for campo, idx in header_map.items()
                            }

                            try:
                                nome = _texto_limpo(dados.get("nome"))
                                if not nome:
                                    raise ValueError("Nome nÃ£o informado.")
                                cpf = _apenas_digitos(_texto_limpo(dados.get("cpfCnpj")))
                                if not cpf:
                                    raise ValueError("CPF/CNPJ nÃ£o informado.")
                                valor_nominal = _parse_decimal(dados.get("valorNominal"))
                                dia_venc = _parse_dia_vencimento(dados.get("dataVencimento"))
                            except (ValueError, InvalidOperation) as exc:
                                erros.append(f"Linha {linha_idx}: {exc}")
                                continue

                            defaults = {
                                "nome": nome,
                                "valorNominal": valor_nominal,
                                "dataVencimento": dia_venc,
                                "email": _texto_limpo(dados.get("email")),
                                "ddd": _texto_limpo(dados.get("ddd")),
                                "telefone": _texto_limpo(dados.get("telefone")),
                                "endereco": _texto_limpo(dados.get("endereco")),
                                "numero": _texto_limpo(dados.get("numero")),
                                "complemento": _texto_limpo(dados.get("complemento")),
                                "bairro": _texto_limpo(dados.get("bairro")),
                                "cidade": _texto_limpo(dados.get("cidade")),
                                "uf": _texto_limpo(dados.get("uf")).upper(),
                                "cep": _texto_limpo(dados.get("cep")),
                            }

                            cliente, criado = Cliente.objects.update_or_create(
                                cpfCnpj=cpf,
                                defaults=defaults,
                            )

                            if criado:
                                criados += 1
                            else:
                                atualizados += 1

                        if criados or atualizados:
                            mensagens = []
                            if criados:
                                mensagens.append(f"{criados} cliente(s) novo(s)")
                            if atualizados:
                                mensagens.append(f"{atualizados} cliente(s) atualizado(s)")
                            resumo = ", ".join(mensagens)
                            messages.success(
                                request,
                                f"ImportaÃ§Ã£o concluÃ­da com sucesso: {resumo}.",
                            )
                        else:
                            messages.info(
                                request,
                                "Nenhum cliente foi criado ou atualizado. Verifique os dados da planilha.",
                            )

                        if erros:
                            resumo_erros = "; ".join(erros[:5])
                            if len(erros) > 5:
                                resumo_erros += f"; ... (+{len(erros) - 5} linha(s) com erro)"
                            messages.warning(
                                request,
                                f"Algumas linhas foram ignoradas: {resumo_erros}",
                            )

                        return redirect("clientes_list")
            finally:
                workbook.close()

    return render(request, "billing/cliente_import.html", {"form": form})


@login_required
def cliente_import_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(
        [
            "Nome",
            "CPF/CNPJ",
            "Valor nominal",
            "dataVencimento",
            "E-mail",
            "DDD",
            "Telefone",
            "Endereco",
            "Numero",
            "Complemento",
            "Bairro",
            "Cidade",
            "UF",
            "CEP",
        ]
    )
    ws.append(
        [
            "Empresa Exemplo Ltda",
            "12.345.678/0001-90",
            199.9,
            10,
            "contato@exemplo.com",
            "11",
            "99999-1111",
            "Rua das Flores",
            "123",
            "Sala 12",
            "Centro",
            "Sao Paulo",
            "SP",
            "01000-000",
        ]
    )
    ws.append(
        [
            "Cliente Pessoa FÃ­sica",
            "123.456.789-00",
            89.5,
            25,
            "cliente@email.com",
            "21",
            "98888-2222",
            "Av. Atlantica",
            "456",
            "",
            "Copacabana",
            "Rio de Janeiro",
            "RJ",
            "22010-000",
        ]
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="clientes_modelo.xlsx"'
    wb.save(response)
    return response


@login_required
def cliente_create(request):
    form = ClienteForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Cliente cadastrado com sucesso.")
        return redirect("clientes_list")
    return render(request, "billing/cliente_form.html", {"form": form, "titulo": "Novo cliente"})


@login_required
def cliente_update(request, cliente_id: int):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    form = ClienteForm(request.POST or None, instance=cliente)
    if form.is_valid():
        form.save()
        messages.success(request, "Cliente atualizado com sucesso.")
        return redirect("clientes_list")
    return render(request, "billing/cliente_form.html", {"form": form, "titulo": f"Editar {cliente.nome}"})


@login_required
def cliente_delete(request, cliente_id: int):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if request.method == "POST":
        cliente.delete()
        messages.success(request, "Cliente removido.")
        return redirect("clientes_list")
    return render(request, "billing/cliente_confirm_delete.html", {"cliente": cliente})






def _aplicar_filtros_boletos(request, queryset, *, incluir_status=True):
    """Aplica filtros compartilhados e retorna o queryset filtrado mais o contexto."""
    hoje = timezone.localdate()
    mes_param_raw = request.GET.get("mes") if "mes" in request.GET else str(hoje.month)
    ano_param_raw = request.GET.get("ano") if "ano" in request.GET else str(hoje.year)

    mes_param = (mes_param_raw or "").strip()
    ano_param = (ano_param_raw or "").strip()
    dia_param = request.GET.get("dia", "").strip()
    nome_param = request.GET.get("nome", "").strip()

    boletos_filtrados = queryset

    mes_selecionado = ""
    if mes_param:
        try:
            mes_valor = int(mes_param)
        except ValueError:
            mes_valor = None
        if mes_valor and 1 <= mes_valor <= 12:
            boletos_filtrados = boletos_filtrados.filter(competencia_mes=mes_valor)
            mes_selecionado = str(mes_valor)

    ano_selecionado = ""
    if ano_param:
        try:
            ano_valor = int(ano_param)
        except ValueError:
            ano_valor = None
        if ano_valor:
            boletos_filtrados = boletos_filtrados.filter(competencia_ano=ano_valor)
            ano_selecionado = str(ano_valor)

    status_opcoes: List[Dict[str, str]] = []
    status_selecionado = ""
    if incluir_status:
        if "status" in request.GET:
            status_param_raw = request.GET.get("status", "")
        else:
            status_param_raw = Boleto.STATUS_EMITIDO
        status_param = (status_param_raw or "").strip()

        status_choices_map = dict(Boleto.STATUS_CHOICES)
        status_choices_visiveis = sorted(
            [
                (value, label)
                for value, label in Boleto.STATUS_CHOICES
                if value != "novo"
            ],
            key=lambda item: item[1],
        )
        status_opcoes = [{"value": "", "label": "Todos"}] + [
            {"value": value, "label": label} for value, label in status_choices_visiveis
        ]
        status_opcoes.append({"value": "pago_pix", "label": "Pago (PIX)"})
        status_choices_map["pago_pix"] = "Pago (PIX)"

        if status_param == "pago_pix":
            boletos_filtrados = boletos_filtrados.filter(status=Boleto.STATUS_PAGO, forma_pagamento="pix")
            status_selecionado = status_param
        elif status_param and status_param in status_choices_map:
            boletos_filtrados = boletos_filtrados.filter(status=status_param)
            status_selecionado = status_param
    else:
        status_selecionado = (request.GET.get("status", "") or "").strip()

    dia_selecionado = ""
    if dia_param:
        try:
            dia_valor = int(dia_param)
        except ValueError:
            dia_valor = None
        if dia_valor and 1 <= dia_valor <= 31:
            boletos_filtrados = boletos_filtrados.filter(data_vencimento__day=dia_valor)
            dia_selecionado = str(dia_valor)

    nome_selecionado = ""
    if nome_param:
        boletos_filtrados = boletos_filtrados.filter(cliente__nome__icontains=nome_param)
        nome_selecionado = nome_param

    anos_disponiveis = list(
        Boleto.objects.order_by("-competencia_ano")
        .values_list("competencia_ano", flat=True)
        .distinct()
    )

    dias_disponiveis = (
        Boleto.objects.annotate(dia=ExtractDay("data_vencimento"))
        .values_list("dia", flat=True)
        .order_by("dia")
        .distinct()
    )

    meses_contexto = [{"value": "", "label": "Todos"}] + [
        {"value": str(valor), "label": nome} for valor, nome in MESES_CHOICES
    ]

    dias_contexto = [{"value": "", "label": "Todos"}] + [
        {"value": str(dia), "label": f"Dia {int(dia):02d}"}
        for dia in dias_disponiveis
        if dia is not None
    ]

    contexto_filtros = {
        "meses": meses_contexto,
        "anos": [str(ano) for ano in anos_disponiveis],
        "dias": dias_contexto,
        "mes_selecionado": mes_selecionado,
        "ano_selecionado": ano_selecionado,
        "dia_selecionado": dia_selecionado,
        "status_opcoes": status_opcoes,
        "status_selecionado": status_selecionado,
        "nome_selecionado": nome_selecionado,
    }

    return boletos_filtrados, contexto_filtros


@login_required
def boletos_list(request):
    boletos_queryset = Boleto.objects.select_related("cliente")
    boletos_filtrados, contexto_filtros = _aplicar_filtros_boletos(request, boletos_queryset)
    boletos = boletos_filtrados.order_by("cliente__nome", "-criado_em")

    context = {
        "boletos": boletos,
        **contexto_filtros,
    }
    return render(request, "billing/boletos_list.html", context)


@login_required
def boleto_create(request):
    form = BoletoForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        boleto = form.save()
        messages.success(request, f"Boleto criado para {boleto.cliente.nome}.")
        return redirect("boletos_list")
    return render(request, "billing/boleto_form.html", {"form": form, "titulo": "Novo boleto"})


@login_required
def boleto_update(request, boleto_id: int):
    boleto = get_object_or_404(Boleto, id=boleto_id)
    form = BoletoForm(request.POST or None, request.FILES or None, instance=boleto)
    if form.is_valid():
        boleto = form.save()
        messages.success(request, f"Boleto atualizado para {boleto.cliente.nome}.")
        return redirect("boletos_list")
    return render(request, "billing/boleto_form.html", {"form": form, "titulo": f"Editar boleto #{boleto.id}"})


@login_required
def boleto_delete(request, boleto_id: int):
    boleto = get_object_or_404(Boleto, id=boleto_id)
    if request.method == "POST":
        boleto.delete()
        messages.success(request, "Boleto removido.")
        return redirect("boletos_list")
    return render(request, "billing/boleto_confirm_delete.html", {"boleto": boleto})


@login_required
def gerar_boletos(request):
    if request.method == "POST":
        form = SelecionarClientesForm(request.POST)
    else:
        get_data = request.GET if request.GET else None
        form = SelecionarClientesForm(get_data)
    if request.method == "POST" and form.is_valid():
        ano = form.cleaned_data["ano"]
        mes = form.cleaned_data["mes"]
        clientes_escolhidos = form.cleaned_data["clientes"]
        if clientes_escolhidos:
            clientes = list(clientes_escolhidos)
        else:
            clientes = list(form.filtered_clientes)

        if not clientes:
            messages.info(request, "Nenhum cliente disponivel para o filtro selecionado.")
            return render(request, "billing/gerar_boletos.html", {"form": form})
        inter = InterService()

        with transaction.atomic():
            for cli in clientes:
                # Calcula data de vencimento (ajustando para ÃƒÂƒÃ†Â’ÃƒÂ†Ã¢Â€Â™ÃƒÂƒÃ¢Â€ÂšÃƒÂ‚Ã‚Âºltimo dia do mÃƒÂƒÃ†Â’ÃƒÂ†Ã¢Â€Â™ÃƒÂƒÃ¢Â€ÂšÃƒÂ‚Ã‚Âªs, se necessÃƒÂƒÃ†Â’ÃƒÂ†Ã¢Â€Â™ÃƒÂƒÃ¢Â€ÂšÃƒÂ‚Ã‚Â¡rio)
                last_day = calendar.monthrange(ano, mes)[1]
                dia = min(cli.dataVencimento, last_day)
                data_venc = dt.date(ano, mes, dia)

                # Evita duplicidade da mesma competÃƒÂƒÃ†Â’ÃƒÂ†Ã¢Â€Â™ÃƒÂƒÃ¢Â€ÂšÃƒÂ‚Ã‚Âªncia
                boleto, created = Boleto.objects.get_or_create(
                    cliente=cli, competencia_ano=ano, competencia_mes=mes,
                    defaults={
                        "data_vencimento": data_venc,
                        "valor": cli.valorNominal,
                    }
                )
                if not created:
                    if boleto.status == Boleto.STATUS_CANCELADO:
                        _preparar_boleto_para_reemissao(boleto)
                        boleto.data_vencimento = data_venc
                        boleto.valor = cli.valorNominal
                    else:
                        messages.info(request, f"Boleto jÃ¡ existia: {cli.nome} {mes:02d}/{ano}")
                        continue

                # Monta dict no formato esperado pelo serviÃƒÂƒÃ†Â’ÃƒÂ†Ã¢Â€Â™ÃƒÂƒÃ¢Â€ÂšÃƒÂ‚Ã‚Â§o (Banco Inter)
                cli_dict = {
                    "valorNominal": float(cli.valorNominal),
                    "nome": cli.nome,
                    "cpfCnpj": cli.cpfCnpj,
                    "email": cli.email,
                    "ddd": DEFAULT_BOLETO_DDD,
                    "telefone": DEFAULT_BOLETO_TELEFONE,
                    "endereco": cli.endereco,
                    "numero": cli.numero,
                    "complemento": cli.complemento,
                    "bairro": cli.bairro,
                    "cidade": cli.cidade,
                    "uf": cli.uf,
                    "cep": cli.cep,
                }
                try:
                    result = inter.emitir_boleto(cli_dict, data_venc)
                    boleto.nosso_numero = result.get("nossoNumero","")
                    boleto.linha_digitavel = result.get("linhaDigitavel","")
                    boleto.codigo_barras = result.get("codigoBarras","")
                    boleto.tx_id = result.get("txId","")
                    boleto.codigo_solicitacao = result.get("codigoSolicitacao","")
                    boleto.status = Boleto.STATUS_EMITIDO
                    boleto.save()

                except Exception as e:
                    boleto.status = Boleto.STATUS_ERRO
                    boleto.erro_msg = str(e)
                    boleto.save()
                    messages.error(request, f"Erro ao emitir boleto de {cli.nome}: {e}")

            messages.success(request, "Processo de emissao finalizado.")
        return redirect("boletos_list")

    return render(request, "billing/gerar_boletos.html", {"form": form})

@login_required
def baixar_pdf_view(request, boleto_id: int):
    boleto = get_object_or_404(Boleto, id=boleto_id)
    inter = InterService()
    pdf_bytes = _buscar_pdf_bytes(inter, boleto)
    if not pdf_bytes:
        messages.info(
            request,
            "PDF ainda nao disponivel na API do Banco Inter. Tente novamente em alguns instantes.",
        )
        return redirect("boletos_list")

    if not _pdf_existe_localmente(boleto):
        filename = Path(boleto.pdf.name).name if boleto.pdf else _arquivo_pdf_nome(boleto)
        boleto.pdf.save(filename, ContentFile(pdf_bytes))
        boleto.save(update_fields=["pdf"])
    if not boleto.codigo_barras:
        _atualizar_codigo_barras_via_inter(inter, boleto)

    stored_name = Path(boleto.pdf.name).name if boleto.pdf else _arquivo_pdf_nome(boleto)
    return FileResponse(
        boleto.pdf.open("rb"),
        as_attachment=True,
        filename=stored_name,
    )


@login_required
def baixar_pdf_lote(request):
    if request.method != "POST":
        messages.info(request, "Selecione os boletos desejados e use o botao de download.")
        return redirect("boletos_list")

    ids = request.POST.getlist("boletos")
    if not ids:
        messages.info(request, "Selecione ao menos um boleto para baixar.")
        return redirect("boletos_list")

    boletos = list(Boleto.objects.filter(id__in=ids).select_related("cliente"))
    if not boletos:
        messages.error(request, "Nenhum boleto encontrado para os identificadores informados.")
        return redirect("boletos_list")

    inter = InterService()
    buffer = io.BytesIO()
    erros: List[str] = []
    nomes_utilizados: Set[str] = set()
    sucesso = 0

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_stream:
        for boleto in boletos:
            pdf_bytes = _buscar_pdf_bytes(inter, boleto)
            if not pdf_bytes:
                erros.append(f"Boleto {boleto.id} - {boleto.cliente.nome}")
                continue

            sucesso += 1
            if not _pdf_existe_localmente(boleto):
                filename = Path(boleto.pdf.name).name if boleto.pdf else _arquivo_pdf_nome(boleto)
                boleto.pdf.save(filename, ContentFile(pdf_bytes))
                boleto.save(update_fields=["pdf"])
            if not boleto.codigo_barras:
                _atualizar_codigo_barras_via_inter(inter, boleto)

            stored_name = Path(boleto.pdf.name).name if boleto.pdf else _arquivo_pdf_nome(boleto)
            nome_zip = stored_name
            base_name = Path(stored_name).stem or f"boleto_{boleto.id}"
            extension = Path(stored_name).suffix or ".pdf"
            contador = 1
            while nome_zip in nomes_utilizados:
                nome_zip = f"{base_name}_{contador}{extension}"
                contador += 1
            nomes_utilizados.add(nome_zip)
            zip_stream.writestr(nome_zip, pdf_bytes)

        if erros:
            conteudo_erros = "Nao foi possivel obter o PDF dos seguintes boletos:\n" + "\n".join(erros)
            zip_stream.writestr("boletos_com_erro.txt", conteudo_erros)

    if sucesso == 0:
        messages.error(request, "Nao foi possivel baixar o PDF de nenhum boleto selecionado.")
        return redirect("boletos_list")

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = "attachment; filename=boletos_selecionados.zip"
    return response

@login_required
def marcar_pago(request, boleto_id: int):
    boleto = get_object_or_404(Boleto, id=boleto_id)
    _registrar_pagamento_manual(boleto, forma_pagamento="")
    messages.success(request, "Baixa registrada no sistema.")
    return redirect("boletos_list")


@login_required
def marcar_pago_pix(request, boleto_id: int):
    boleto = get_object_or_404(Boleto, id=boleto_id)
    _registrar_pagamento_manual(boleto, forma_pagamento="pix")
    messages.success(request, "Baixa registrada via PIX.")
    return redirect("boletos_list")


@login_required
def marcar_pago_dinheiro(request, boleto_id: int):
    boleto = get_object_or_404(Boleto, id=boleto_id)
    _registrar_pagamento_manual(boleto, forma_pagamento="dinheiro")
    messages.success(request, "Baixa registrada como recebimento em dinheiro.")
    return redirect("boletos_list")


def _registrar_pagamento_manual(
    boleto: Boleto, forma_pagamento: str, data_pagamento: Optional[dt.date] = None
) -> None:
    boleto.status = Boleto.STATUS_PAGO
    boleto.forma_pagamento = forma_pagamento or ""
    boleto.data_pagamento = data_pagamento or dt.date.today()
    boleto.save(update_fields=["status", "forma_pagamento", "data_pagamento"])


@login_required
def cancelar_boleto(request, boleto_id: int):
    boleto = get_object_or_404(Boleto, id=boleto_id)
    inter = InterService()
    try:
        resultado = inter.cancelar_boleto(
            codigo_solicitacao=boleto.codigo_solicitacao or "",
            nosso_numero=boleto.nosso_numero or "",
        )
    except Exception as exc:  # noqa: BLE001 - queremos exibir o motivo ao usuÃƒÂƒÃ†Â’ÃƒÂ†Ã¢Â€Â™ÃƒÂƒÃ¢Â€ÂšÃƒÂ‚Ã‚Â¡rio
        boleto.erro_msg = str(exc)
        boleto.save(update_fields=["erro_msg"])
        messages.error(request, f"Falha ao cancelar via API: {exc}")
    else:
        boleto.status = Boleto.STATUS_CANCELADO
        boleto.erro_msg = ""
        boleto.save(update_fields=["status", "erro_msg"])
        situacao = resultado.get("situacao") or resultado.get("status")
        if situacao:
            messages.success(
                request,
                f"Cobran\u00e7a cancelada. Situa\u00e7\u00e3o informada pelo Inter: {situacao}",
            )
        else:
            messages.success(request, "Cobran\u00e7a cancelada com sucesso no Inter.")
    return redirect("boletos_list")





def _file_info(arquivo) -> Dict[str, Any]:
    if not arquivo:
        return {"name": "", "path": "", "exists": False}
    try:
        caminho = Path(arquivo.path)
    except (ValueError, OSError):
        caminho = None
    return {
        "name": arquivo.name,
        "path": str(caminho) if caminho else "",
        "exists": caminho.exists() if caminho else False,
    }


@login_required
@require_http_methods(["GET", "POST"])
def config_inter(request):
    config = InterConfig.get_solo()
    form = InterConfigForm(request.POST or None, request.FILES or None, instance=config)

    cert_info = _file_info(config.cert_file)
    key_info = _file_info(config.key_file)

    if request.method == "POST" and form.is_valid():
        config = form.save()
        cert_info = _file_info(config.cert_file)
        key_info = _file_info(config.key_file)
        messages.success(request, "Configuracao do Banco Inter atualizada.")
        return redirect("config_inter")

    credentials_dir = Path(settings.MEDIA_ROOT) / "inter_credentials"

    context = {
        "form": form,
        "config": config,
        "credentials_dir": credentials_dir,
        "cert_info": cert_info,
        "key_info": key_info,
    }
    return render(request, "billing/inter_config.html", context)



@login_required
@require_http_methods(["GET", "POST"])
def enviar_boletos_whatsapp(request):
    config = WhatsappConfig.get_solo()
    mensagem_form = WhatsappMensagemForm(instance=config)


    boletos_queryset = Boleto.objects.select_related("cliente")
    boletos_filtrados, contexto_filtros = _aplicar_filtros_boletos(
        request,
        boletos_queryset,
        incluir_status=False,
    )
    boletos_filtrados = boletos_filtrados.filter(status=Boleto.STATUS_EMITIDO)
    status_envio_opcoes = [
        {"value": "", "label": "Todos"},
        {"value": "enviado", "label": "Enviado"},
        {"value": "a_enviar", "label": "A enviar"},
        {"value": "erro", "label": "Erro"},
    ]
    status_envio_param = (request.GET.get("status", "") or "").strip().lower()
    valores_status_validos = {op["value"] for op in status_envio_opcoes if op["value"]}
    if status_envio_param and status_envio_param not in valores_status_validos:
        status_envio_param = ""
    contexto_filtros.update(
        {
            "status_opcoes": status_envio_opcoes,
            "status_selecionado": status_envio_param,
        }
    )
    boletos_queryset = boletos_filtrados.order_by("data_vencimento", "id")
    boletos = list(boletos_queryset)

    session_status_map = dict(request.session.get("boletos_envio_status", {}))
    session_detail_map = dict(request.session.get("boletos_envio_detail", {}))
    chaves_atuais = {str(boleto.id) for boleto in boletos}
    session_status_map = {chave: valor for chave, valor in session_status_map.items() if chave in chaves_atuais}
    session_detail_map = {chave: valor for chave, valor in session_detail_map.items() if chave in chaves_atuais}
    request.session["boletos_envio_status"] = session_status_map
    request.session["boletos_envio_detail"] = session_detail_map

    def _detalhe_envio(res: Optional[Dict[str, Any]]) -> str:
        if not res:
            return ""
        if res.get("ok"):
            return ""
        mensagem = res.get("error")
        if mensagem:
            return mensagem
        for step in reversed(res.get("steps", [])):
            payload = step.get("payload")
            if isinstance(payload, dict):
                for chave in ("message", "error", "details", "status"):
                    if payload.get(chave):
                        return str(payload[chave])
            elif isinstance(payload, str):
                return payload
        return ""

    def _status_envio_slug(label: Optional[str]) -> str:
        texto = (label or "").strip().lower()
        if "erro" in texto:
            return "erro"
        if "enviado" in texto:
            return "enviado"
        return "a_enviar"


    resultados: List[Dict[str, Any]] = []
    alvo_ids_raw = request.POST.getlist("boleto_id") if request.method == "POST" else []

    if request.method == "POST":
        acao = request.POST.get("acao")
        if acao == "atualizar_mensagem":
            mensagem_form = WhatsappMensagemForm(request.POST, instance=config)
            if mensagem_form.is_valid():
                mensagem_form.save()
                messages.success(request, "Mensagem de envio atualizada com sucesso.")
                return redirect("enviar_boletos_whatsapp")
            messages.error(request, "Corrija os erros no formulario de configuracao.")
        else:
            if alvo_ids_raw:
                boletos_alvo = [boleto for boleto in boletos if str(boleto.id) in alvo_ids_raw]
            else:
                boletos_alvo = boletos

            boletos_validos: List[Boleto] = []
            boletos_invalidos: List[Boleto] = []
            for boleto in boletos_alvo:
                if boleto.status != Boleto.STATUS_EMITIDO:
                    boletos_invalidos.append(boleto)
                    continue
                boletos_validos.append(boleto)

            if boletos_invalidos:
                ignorados_nomes = ", ".join(boleto.cliente.nome for boleto in boletos_invalidos)
                messages.warning(
                    request,
                    f"Alguns boletos foram ignorados por nao estarem com status emitido: {ignorados_nomes}.",
                )

            for boleto in boletos_validos:
                resultado = dispatch_boleto_via_whatsapp(
                    boleto,
                    saudacao_template=config.saudacao_template,
                )
                resultados.append(resultado)

                detalhe_envio = _detalhe_envio(resultado)
                if resultado.get("ok"):
                    novo_status = Boleto.WHATSAPP_STATUS_ENVIADO
                else:
                    novo_status = Boleto.WHATSAPP_STATUS_ERRO
                boleto.whatsapp_status = novo_status
                boleto.whatsapp_status_detail = detalhe_envio or ""
                boleto.whatsapp_status_updated_at = timezone.now()
                boleto.save(
                    update_fields=[
                        "whatsapp_status",
                        "whatsapp_status_detail",
                        "whatsapp_status_updated_at",
                    ]
                )

                chave_boleto = str(boleto.id)
                session_status_map[chave_boleto] = boleto.get_whatsapp_status_display()
                session_detail_map[chave_boleto] = boleto.whatsapp_status_detail

            enviados_sucesso = sum(1 for resultado in resultados if resultado.get("ok"))
            erros_envio = sum(1 for resultado in resultados if not resultado.get("ok"))
            if enviados_sucesso:
                messages.success(request, f"{enviados_sucesso} boleto(s) enviado(s) com sucesso via WhatsApp.")
            if erros_envio:
                messages.error(request, f"{erros_envio} envio(s) falharam. Verifique os detalhes na listagem abaixo.")

    for resultado in resultados:
        boleto_id = resultado.get("boleto_id")
        if not boleto_id:
            continue
        chave = str(boleto_id)
        session_status_map[chave] = "Enviado" if resultado.get("ok") else "Erro"
        session_detail_map[chave] = _detalhe_envio(resultado)

    if resultados:
        request.session["boletos_envio_status"] = session_status_map
        request.session["boletos_envio_detail"] = session_detail_map

    status_map: Dict[int, str] = {}
    for boleto in boletos:
        chave = str(boleto.id)
        status_label = session_status_map.get(chave)
        if not status_label:
            if boleto.whatsapp_status:
                status_label = boleto.get_whatsapp_status_display()
            else:
                status_label = "A enviar"
        status_map[boleto.id] = status_label

    linhas_boletos: List[Tuple[Dict[str, Any], str]] = []
    for boleto in boletos:
        cliente = boleto.cliente
        telefone_whatsapp = format_whatsapp_phone(cliente)
        telefone_display = telefone_whatsapp.split("@")[0] if telefone_whatsapp else ""
        bloqueios: List[str] = []
        if boleto.status != Boleto.STATUS_EMITIDO:
            descricao_status = boleto.get_status_display()
            bloqueios.append(f"Status atual impede envio: {descricao_status}.")
        pdf_disponivel = _pdf_existe_localmente(boleto)
        if not telefone_whatsapp:
            bloqueios.append("Telefone do cliente invalido ou ausente.")
        if not pdf_disponivel:
            bloqueios.append("PDF do boleto ainda nao foi baixado.")
        status_envio_label = status_map.get(boleto.id, "A enviar")
        status_envio_slug = _status_envio_slug(status_envio_label)
        linha = {
            "id": boleto.id,
            "cliente": cliente.nome,
            "competencia": f"{boleto.competencia_mes:02d}/{boleto.competencia_ano}",
            "valor": boleto.valor,
            "vencimento": boleto.data_vencimento,
            "telefone": telefone_display,
            "telefone_whatsapp": telefone_whatsapp,
            "telefone_bruto": f"{cliente.ddd or ''}{cliente.telefone or ''}",
            "codigo_barras": boleto.codigo_barras or boleto.linha_digitavel or "",
            "pdf_url": boleto.pdf.url if pdf_disponivel else "",
            "pdf_disponivel": pdf_disponivel,
            "status_envio": status_envio_label,
            "pode_enviar": not bloqueios,
            "bloqueios": bloqueios,
            "detalhe_envio": session_detail_map.get(
                str(boleto.id),
                boleto.whatsapp_status_detail or "",
            ),
        }
        linhas_boletos.append((linha, status_envio_slug))

    tabela_boletos = [linha for linha, slug in linhas_boletos if not status_envio_param or slug == status_envio_param]

    total = len(tabela_boletos)
    enviados = sum(1 for item in tabela_boletos if item["status_envio"] == "Enviado")
    erros = sum(1 for item in tabela_boletos if item["status_envio"] == "Erro")
    pendentes = total - enviados - erros

    context = {
        **contexto_filtros,
        "boletos": tabela_boletos,
        "total": total,
        "total_enviados": enviados,
        "total_erros": erros,
        "total_pendentes": pendentes,
        "resultados": resultados,
        "alvo_ids": alvo_ids_raw,
        "mensagem_form": mensagem_form,
        "mensagem_config": config,
    }
    return render(request, "billing/enviar_boletos_whatsapp.html", context)
